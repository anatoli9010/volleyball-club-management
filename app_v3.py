# app_v3.py
import os
import re
import time
import threading
from pathlib import Path
from typing import Optional
from datetime import datetime, date, timezone, timedelta
from functools import wraps
import logging

from flask import (
    Flask, render_template, redirect, url_for, flash, request, abort, jsonify
)
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from apscheduler.schedulers.background import BackgroundScheduler
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

import requests
import smtplib
from email.message import EmailMessage
import pandas as pd
from dotenv import load_dotenv
from sqlalchemy import case, extract

# Telegram imports
from telegram import Update
from telegram.ext import Application

# -------------------- Load .env --------------------
BASE_DIR = Path(__file__).resolve().parent
ENV_PATH = BASE_DIR / '.env'
if ENV_PATH.exists():
    load_dotenv(dotenv_path=str(ENV_PATH))

# -------------------- Config --------------------
INSTANCE_DIR = BASE_DIR / 'instance'
TEMPLATES_DIR = BASE_DIR / 'templates'
DB_PATH = INSTANCE_DIR / 'trenera.db'
UPLOAD_FOLDER = BASE_DIR / 'uploads'
ALLOWED_EXT = {'csv'}  # CSV-only

os.makedirs(INSTANCE_DIR, exist_ok=True)
os.makedirs(TEMPLATES_DIR, exist_ok=True)
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

app = Flask(__name__, template_folder=str(TEMPLATES_DIR))
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret')

# Database configuration - PostgreSQL for production, SQLite for development
DATABASE_URL = os.environ.get('DATABASE_URL')
if DATABASE_URL:
    # Production - PostgreSQL (Render, Railway, Supabase, etc.)
    if DATABASE_URL.startswith('postgres://'):
        DATABASE_URL = DATABASE_URL.replace('postgres://', 'postgresql://', 1)
    app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL
else:
    # Development - SQLite
    app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{DB_PATH}'

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = str(UPLOAD_FOLDER)

# In-memory mapping (can be populated by CSV import or bot binds)
phone_to_telegram = {}

# Telegram and SMTP config read from .env
TELEGRAM_BOT_TOKEN = os.environ.get('TELEGRAM_BOT_TOKEN', '').strip()
SMTP_HOST = os.environ.get('SMTP_HOST', '')
SMTP_PORT = int(os.environ.get('SMTP_PORT', '0')) if os.environ.get('SMTP_PORT') else None
SMTP_USER = os.environ.get('SMTP_USER', '')
SMTP_PASS = os.environ.get('SMTP_PASS', '')

# -------------------- Logging --------------------
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger('trenera')

# -------------------- Extensions --------------------
db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'

# -------------------- Telegram webhook (simple) --------------------
import json

def normalize_phone_for_match(ph):
    """Връща последните 9 цифри (за сравнение на бг номера)."""
    if not ph:
        return None
    digits = re.sub(r'\D', '', str(ph))
    # keep last 9 digits (e.g. 888123456)
    return digits[-9:] if len(digits) >= 9 else digits

def send_telegram(chat_id, text):
    """Изпраща съобщение към Telegram bot API (проста реализация)."""
    if not TELEGRAM_BOT_TOKEN or not chat_id:
        logger.info("send_telegram - липсва token или chat_id")
        return
    try:
        requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage",
            json={"chat_id": str(chat_id), "text": text, "disable_web_page_preview": True}
        )
    except Exception:
        logger.exception("send_telegram failed")

@app.route(f"/webhook/{TELEGRAM_BOT_TOKEN}", methods=["POST"])
def telegram_webhook():
    update = request.get_json(force=True)

    if "message" in update:
        chat_id = str(update["message"]["chat"]["id"])
        text = update["message"].get("text", "").strip()

        logger.info(f"[TELEGRAM] Message from chat_id={chat_id}: {text}")

        # Handle /start command
        if text.lower() in ("/start", "старт"):
            send_telegram(
                chat_id,
                "Добре дошли в системата Trenera! 📲\n"
                "Моля, въведете вашия телефонен номер (пример: 0888123456), за да активирате известия."
            )
            return "OK", 200

        # Normalize phone number
        def normalize_phone(phone_str):
            if not phone_str:
                return None
            digits = re.sub(r'\D', '', phone_str)
            if digits.startswith('0') and len(digits) == 10:
                return '359' + digits[1:]
            elif digits.startswith('359') and len(digits) == 11:
                return digits
            elif digits.startswith('00359') and len(digits) == 12:
                return digits[2:]
            return digits[-9:]  # fallback

        phone_number = normalize_phone(text)
        logger.info(f"[TELEGRAM] Normalized phone: {phone_number} from input: {text}")

        if not phone_number or len(phone_number) < 9:
            send_telegram(chat_id, "❌ Невалиден формат на телефон. Пример: 0888123456")
            return "OK", 200

        # Search for player by parent_phone or player_phone
        players = Player.query.all()
        matched_players = []

        for player in players:
            parent_normalized = normalize_phone(player.parent_phone)
            player_normalized = normalize_phone(player.player_phone)

            # DEBUG LOG
            logger.info(
                f"[MATCH DEBUG] Player {player.id} {player.full_name} "
                f"parent_phone={player.parent_phone} -> norm={parent_normalized} "
                f"player_phone={player.player_phone} -> norm={player_normalized} "
                f"input_norm={phone_number}"
            )

            if (parent_normalized and parent_normalized.endswith(phone_number[-9:])) or \
               (player_normalized and player_normalized.endswith(phone_number[-9:])):
                matched_players.append(player)

        if matched_players:
            logger.info(f"[DB] Found {len(matched_players)} matching players")

            for player in matched_players:
                logger.info(f"[DB] Updating player {player.id} {player.full_name}")
                logger.info(f"[DB] Old telegram_id: {player.parent_telegram_id}")
                player.parent_telegram_id = chat_id
                db.session.add(player)

            try:
                db.session.commit()
                logger.info("[DB] Commit executed")

                for player in matched_players:
                    check = Player.query.get(player.id)
                    logger.info(f"[DB CHECK] After commit -> player {check.id} telegram_id={check.parent_telegram_id}")

                send_telegram(
                    chat_id,
                    f"✅ Вашият номер беше регистриран успешно за {len(matched_players)} състезател(и).\n\n"
                    "Отсега нататък ще получавате известия за:\n"
                    "• 📅 Присъствие/отсъствие от тренировки\n"
                    "• 💰 Плащане на месечна такса\n"
                    "• ⏰ Напомняния за плащане"
                )
            except Exception as e:
                logger.error(f"[DB] Failed to commit changes: {str(e)}")
                send_telegram(chat_id, "❌ Възникна грешка при регистрацията. Моля, опитайте отново.")
        else:
            logger.warning(f"[DB] No player found with phone ending with: {phone_number[-9:]}")
            send_telegram(
                chat_id,
                "❌ Този номер не е намерен в системата.\n"
                "Моля, свържете се с треньора, за да ви добави."
            )

    return "OK", 200




def maybe_set_webhook():
    """Опционално задава webhook при старт ако сме в Render и имаме hostname + token."""
    try:
        host = os.environ.get('RENDER_EXTERNAL_HOSTNAME') or os.environ.get('EXTERNAL_HOSTNAME')
        if TELEGRAM_BOT_TOKEN and host:
            url = f"https://{host}/webhook/{TELEGRAM_BOT_TOKEN}"
            r = requests.post(f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/setWebhook", json={"url": url}, timeout=10)
            if r.ok:
                logger.info("Telegram webhook set -> %s", url)
            else:
                logger.warning("Failed to set webhook: %s %s", r.status_code, r.text)
    except Exception:
        logger.exception("maybe_set_webhook failed")


# -------------------- Models --------------------

class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    role = db.Column(db.String(20), nullable=False)  # 'admin' or 'trainer'

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

class Team(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    age_group = db.Column(db.String(50), nullable=True)
    gender = db.Column(db.String(10), nullable=True)  # boys/girls
    players = db.relationship('Player', backref='team', lazy=True)

class Player(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    full_name = db.Column(db.String(200), nullable=False)
    birth_date = db.Column(db.Date, nullable=True)
    player_phone = db.Column(db.String(50), nullable=True)
    parent_phone = db.Column(db.String(50), nullable=True)
    parent_telegram_id = db.Column(db.String(80), nullable=True)
    email = db.Column(db.String(120), nullable=True)
    notes = db.Column(db.String(400), nullable=True)
    team_id = db.Column(db.Integer, db.ForeignKey('team.id'), nullable=True)

    payments = db.relationship('Payment', backref='player', lazy=True, cascade='all, delete-orphan')
    attendances = db.relationship('Attendance', backref='player', lazy=True)

class Payment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    player_id = db.Column(db.Integer, db.ForeignKey('player.id'), nullable=False)
    year = db.Column(db.Integer, nullable=False)
    month = db.Column(db.Integer, nullable=False)
    amount = db.Column(db.Float, nullable=False, default=0.0)
    status = db.Column(db.String(20), nullable=False, default='pending')  # pending, paid
    paid_at = db.Column(db.DateTime, nullable=True)
    note = db.Column(db.String(255), nullable=True)

class TrainingSession(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    team_id = db.Column(db.Integer, db.ForeignKey('team.id'), nullable=True)
    date = db.Column(db.Date, nullable=False)
    start_time = db.Column(db.String(20), nullable=True)  # optional "18:00"
    end_time = db.Column(db.String(20), nullable=True)    # optional "19:30"
    notes = db.Column(db.String(400), nullable=True)
    attendances = db.relationship('Attendance', backref='session', lazy=True, cascade='all, delete-orphan')

class Attendance(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    session_id = db.Column(db.Integer, db.ForeignKey('training_session.id'), nullable=False)
    player_id = db.Column(db.Integer, db.ForeignKey('player.id'), nullable=False)
    status = db.Column(db.String(20), nullable=False, default='absent')  # present/absent
    noted_at = db.Column(db.DateTime, nullable=True)

class Season(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    start_date = db.Column(db.Date, nullable=True)
    end_date = db.Column(db.Date, nullable=True)
    is_active = db.Column(db.Boolean, default=False)

class RecurringSlot(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    season_id = db.Column(db.Integer, db.ForeignKey('season.id'), nullable=True)
    team_id = db.Column(db.Integer, db.ForeignKey('team.id'), nullable=True)
    weekday = db.Column(db.Integer, nullable=False)  # 0=Mon ... 6=Sun
    start_time = db.Column(db.String(5), nullable=False)  # HH:MM
    end_time = db.Column(db.String(5), nullable=False)
    venue = db.Column(db.String(50), nullable=True)  # НУПИ / Чавдар / Стадион
    title = db.Column(db.String(120), nullable=True)  # optional label (e.g. Момичета до 12г)

# --- Coaches management ---
class CoachProfile(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    full_name = db.Column(db.String(120), nullable=True)
    phone = db.Column(db.String(50), nullable=True)
    # права (прост модел)
    can_manage_players = db.Column(db.Boolean, default=True)
    can_manage_payments = db.Column(db.Boolean, default=True)
    can_mark_attendance = db.Column(db.Boolean, default=True)
    can_manage_slots = db.Column(db.Boolean, default=False)
    can_manage_tournaments = db.Column(db.Boolean, default=True)
    can_manage_inventory = db.Column(db.Boolean, default=True)

class CoachTeam(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    coach_id = db.Column(db.Integer, db.ForeignKey('coach_profile.id'), nullable=False)
    team_id = db.Column(db.Integer, db.ForeignKey('team.id'), nullable=False)

# -------------------- Auth & Roles --------------------
@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# Цветова схема и нормализация на имената на отбори (консистентна навсякъде)
TEAM_COLORS = {
    'U-12 Ж': '#dc3545',  # червено
    'U-12 М': '#0d6efd',  # синьо
    'U-18 Ж': '#d63384',  # магента/розово
    'U-18 М': '#198754',  # зелено
    'Старша': '#6f42c1',  # лилаво (ако се ползва)
}

def _normalize_team_label(raw: str) -> str:
    """Мапва различни варианти към точните етикети от графика."""
    if not raw:
        return ''
    n = (raw or '').strip().lower()
    # ключови нормализации по кирилица/латиница
    repl = (
        ('u12', 'u-12'),
        ('u 12', 'u-12'),
        ('u18', 'u-18'),
        ('u 18', 'u-18'),
        ('girls', 'ж'),
        ('girls', 'ж'),
        ('boys', 'м'),
        ('момичета', 'ж'),
        ('момчета', 'м'),
        ('мъже', 'м'),
    )
    for a, b in repl:
        n = n.replace(a, b)
    n = n.replace('  ', ' ')
    # точни разпознавания
    if 'u-12' in n and 'ж' in n:
        return 'U-12 Ж'
    if 'u-12' in n and 'м' in n:
        return 'U-12 М'
    if 'u-18' in n and 'ж' in n:
        return 'U-18 Ж'
    if 'u-18' in n and 'м' in n:
        return 'U-18 М'
    if 'старша' in n:
        return 'Старша'
    # вече може да е точно име
    cap = raw.strip()
    if cap in TEAM_COLORS:
        return cap
    return cap

def team_color_for_name(name: str) -> str:
    key = _normalize_team_label(name)
    return TEAM_COLORS.get(key, '#0d6efd')

# направи модела Team достъпен в Jinja шаблоните като 'Team' и функцията team_color
@app.context_processor
def inject_models():
    return dict(Team=Team, team_color=team_color_for_name)
 

def role_required(role):
    def decorator(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            if not current_user.is_authenticated:
                return login_manager.unauthorized()
            if current_user.role != role and current_user.role != 'admin':
                abort(403)
            return fn(*args, **kwargs)
        return wrapper
    return decorator

def coach_permission_required(permission_attr: str):
    def decorator(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            if not current_user.is_authenticated:
                return login_manager.unauthorized()
            if current_user.role == 'admin':
                return fn(*args, **kwargs)
            if current_user.role != 'trainer':
                abort(403)
            cp = CoachProfile.query.filter_by(user_id=current_user.id).first()
            if not cp or not getattr(cp, permission_attr, False):
                abort(403)
            return fn(*args, **kwargs)
        return wrapper
    return decorator

# -------------------- Messaging helpers --------------------
def send_email(to_email, subject, body):
    if not to_email:
        logger.info(f"No email provided; stub send: {subject}")
        return False
    if SMTP_HOST and SMTP_PORT and SMTP_USER and SMTP_PASS:
        try:
            msg = EmailMessage()
            msg['From'] = SMTP_USER
            msg['To'] = to_email
            msg['Subject'] = subject
            msg.set_content(body)
            with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as s:
                s.starttls()
                s.login(SMTP_USER, SMTP_PASS)
                s.send_message(msg)
            logger.info(f"Email sent to {to_email}: {subject}")
            return True
        except Exception as e:
            logger.exception(f"Failed to send email to {to_email}: {e}")
            return False
    else:
        logger.info(f"(Stub) Email to {to_email}: {subject} - {body}")
        return True

def send_telegram(chat_id, message):
    if not chat_id:
        logger.info(f"No chat_id; stub telegram: {message}")
        return False
    if TELEGRAM_BOT_TOKEN:
        try:
            url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
            resp = requests.post(url, json={"chat_id": str(chat_id), "text": message})
            if resp.ok:
                logger.info(f"Telegram sent to {chat_id}")
                return True
            else:
                logger.warning(f"Telegram API failed: {resp.status_code} {resp.text}")
                return False
        except Exception as e:
            logger.exception(f"Error sending telegram: {e}")
            return False
    else:
        logger.info(f"(Stub) Telegram to {chat_id}: {message}")
        return True

def resolve_telegram_id_by_phone(phone):
    if not phone:
        return None
    normalized = phone.strip()
    t = phone_to_telegram.get(normalized)
    if t:
        return t
    p = Player.query.filter_by(parent_phone=normalized).first()
    if p and p.parent_telegram_id:
        return p.parent_telegram_id
    return None

# -------------------- Template creation --------------------
BASIC_TEMPLATES = {}

# Minimal templates (we'll add others if missing)
BASIC_TEMPLATES['base.html'] = """<!doctype html>
<html lang="bg">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Trenera</title>
  <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
</head>
<body>
<nav class="navbar navbar-expand-lg navbar-light bg-light mb-4">
  <div class="container-fluid">
    <a class="navbar-brand" href="{{ url_for('index') }}">🔹 Trenera</a>
    <div class="collapse navbar-collapse">
      <ul class="navbar-nav me-auto mb-2 mb-lg-0">
        {% if current_user.is_authenticated %}
          <li class="nav-item"><a class="nav-link" href="{{ url_for('players') }}">Играчите</a></li>
          <li class="nav-item"><a class="nav-link" href="{{ url_for('payments') }}">Плащания</a></li>
          <li class="nav-item"><a class="nav-link" href="{{ url_for('teams') }}">Отбори</a></li>
          <li class="nav-item"><a class="nav-link" href="{{ url_for('trainings') }}">Тренировки</a></li>
          {% if current_user.role == 'admin' %}<li class="nav-item"><a class="nav-link" href="{{ url_for('admin_panel') }}">Админ</a></li>{% endif %}
        {% endif %}
      </ul>
    </div>
    <div class="d-flex">
      {% if current_user.is_authenticated %}
        <span class="me-2">{{ current_user.username }} ({{ current_user.role }})</span>
        <a class="btn btn-outline-secondary btn-sm me-2" href="{{ url_for('logout') }}">Изход</a>
      {% else %}
        <a class="btn btn-outline-primary btn-sm" href="{{ url_for('login') }}">Вход</a>
      {% endif %}
    </div>
  </div>
</nav>
<div class="container">
  {% with messages = get_flashed_messages(with_categories=true) %}
    {% if messages %}
      {% for cat, msg in messages %}
        <div class="alert alert-{{ 'success' if cat=='success' else 'danger' }} alert-dismissible fade show" role="alert">
          {{ msg }}
          <button type="button" class="btn-close" data-bs-dismiss="alert" aria-label="Close"></button>
        </div>
      {% endfor %}
    {% endif %}
  {% endwith %}
  {% block content %}{% endblock %}
</div>
<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
</body>
</html>"""

BASIC_TEMPLATES['login.html'] = """{% extends 'base.html' %}{% block content %}<div class="row justify-content-center"><div class="col-md-5"><h3 class="mb-3">🔐 Вход</h3><form method="post"><div class="mb-3"><label>Потребителско име</label><input class="form-control" name="username" required></div><div class="mb-3"><label>Парола</label><input class="form-control" name="password" type="password" required></div><button class="btn btn-primary">Вход</button></form></div></div>{% endblock %}"""

# players, player_form, payments, teams, trainings, attendance, stats, admin templates
BASIC_TEMPLATES['index.html'] = """{% extends 'base.html' %}{% block content %}<h2>Добре дошли</h2><p>Изберете действие.</p>{% endblock %}"""

BASIC_TEMPLATES['players.html'] = """{% extends 'base.html' %}{% block content %}
<div class="d-flex justify-content-between align-items-center mb-3">
  <h3>Играчите</h3>
  <a class="btn btn-success" href="{{ url_for('add_player') }}">➕ Добави</a>
</div>
<form class="row g-2 mb-3" method="get">
  <div class="col-auto">
    <input name="q" class="form-control" placeholder="Търси по име..." value="{{ request.args.get('q','') }}">
  </div>
  <div class="col-auto">
    <select name="team_id" class="form-select" onchange="this.form.submit()">
      <option value="">Всички отбори</option>
      {% for t in Team.query.order_by(Team.name).all() %}
        <option value="{{ t.id }}" {% if request.args.get('team_id') and request.args.get('team_id')|int==t.id %}selected{% endif %}>{{ t.name }}</option>
      {% endfor %}
    </select>
  </div>
  <div class="col-auto">
    <button class="btn btn-outline-secondary">Филтър</button>
  </div>
</form>
<div class="row">
{% for p in players %}
  <div class="col-12 col-md-6 col-lg-4">
    <div class="card mb-3">
      <div class="card-body">
        <h5 class="card-title">{{ p.full_name }}</h5>
        <p class="mb-1"><strong>Отбор:</strong> {{ p.team.name if p.team else '—' }}</p>
        <p class="mb-1"><strong>Родител:</strong> {{ p.parent_phone or '—' }}</p>
        <p class="mb-1"><strong>Email:</strong> {{ p.email or '—' }}</p>
        <div class="mt-2">
          <a class="btn btn-sm btn-primary" href="{{ url_for('edit_player', player_id=p.id) }}">✏ Редактирай</a>
          <a class="btn btn-sm btn-outline-secondary" href="{{ url_for('player_payments', player_id=p.id) }}">💳 Плащания</a>
          <a class="btn btn-sm btn-warning" href="{{ url_for('remind_player', player_id=p.id) }}">🔔 Напомни</a>
        </div>
      </div>
    </div>
  </div>
{% endfor %}
</div>
{% endblock %}"""

BASIC_TEMPLATES['player_form.html'] = """{% extends 'base.html' %}{% block content %}
<h3>{{ 'Редактирай' if player else 'Нов състезател' }}</h3>
<form method="post" class="row g-2">
  <div class="col-12"><label>Име</label><input name="full_name" class="form-control" value="{{ player.full_name if player else '' }}" required></div>
  <div class="col-6"><label>Дата ражд.</label><input name="birth_date" type="date" class="form-control" value="{{ player.birth_date if player and player.birth_date else '' }}"></div>
  <div class="col-6"><label>Отбор</label><select name="team_id" class="form-select"><option value=''>-- без отбор --</option>{% for t in teams %}<option value='{{ t.id }}' {% if player and player.team_id==t.id %}selected{% endif %}>{{ t.name }}</option>{% endfor %}</select></div>
  <div class="col-6"><label>Тел. съст.</label><input name="player_phone" class="form-control" value="{{ player.player_phone if player else '' }}"></div>
  <div class="col-6"><label>Тел. родител</label><input name="parent_phone" class="form-control" value="{{ player.parent_phone if player else '' }}"></div>
  <div class="col-6"><label>Email</label><input name="email" class="form-control" value="{{ player.email if player else '' }}"></div>
  <div class="col-6"><label>Telegram ID</label><input name="parent_telegram_id" class="form-control" value="{{ player.parent_telegram_id if player else '' }}"></div>
  <div class="col-12"><label>Бележки</label><textarea name="notes" class="form-control">{{ player.notes if player else '' }}</textarea></div>
  <div class="col-12"><button class="btn btn-primary">Запази</button></div>
</form>
{% endblock %}"""

BASIC_TEMPLATES['payments.html'] = """{% extends 'base.html' %}{% block content %}
<h3>Плащания</h3>
<form class="row g-2 mb-3" method="get">
  <div class="col-auto"><input name="search" class="form-control" placeholder="Търси" value="{{ request.args.get('search','') }}"></div>
  <div class="col-auto"><input name="year" class="form-control" placeholder="Година" value="{{ request.args.get('year','') }}"></div>
  <div class="col-auto"><input name="month" class="form-control" placeholder="Месец" value="{{ request.args.get('month','') }}"></div>
  <div class="col-auto"><button class="btn btn-outline-primary">Филтрирай</button></div>
</form>
<div class="mb-3 text-end">
  <form method="post" action="{{ url_for('remind_all_payments') }}" class="d-inline">
    <button class="btn btn-warning">🔔 Напомни на всички неплатили</button>
  </form>
</div>
<table class="table table-striped">
  <thead><tr><th>Играч</th><th>Година</th><th>Месец</th><th>Сума</th><th>Статус</th><th>Действие</th></tr></thead>
  <tbody>
    {% for pay in payments %}
    <tr>
      <td>{{ pay.player.full_name }}</td><td>{{ pay.year }}</td><td>{{ pay.month }}</td><td>{{ pay.amount }} лв.</td>
      <td>{% if pay.status=='paid' %}<span class="badge bg-success">Платено</span>{% else %}<span class="badge bg-danger">Неплатено</span>{% endif %}</td>
      <td>
        {% if pay.status != 'paid' %}
          <div class="btn-group">
            <button class="btn btn-sm btn-success dropdown-toggle" data-bs-toggle="dropdown">Плати</button>
            <ul class="dropdown-menu">
              {% for m in range(1,13) %}<li><a class="dropdown-item" href="{{ url_for('mark_paid', payment_id=pay.id) }}?month={{ m }}&year={{ pay.year }}">Месец {{ m }}</a></li>{% endfor %}
            </ul>
          </div>
          <form method="post" action="{{ url_for('remind_payment', payment_id=pay.id) }}" style="display:inline"><button class="btn btn-sm btn-outline-warning">🔔</button></form>
        {% endif %}
      </td>
    </tr>
    {% endfor %}
  </tbody>
</table>
{% endblock %}"""

BASIC_TEMPLATES['teams.html'] = """{% extends 'base.html' %}{% block content %}
<div class="d-flex justify-content-between align-items-center mb-3">
  <h3>Отбори</h3>
  {% if current_user.role=='admin' %}<a class="btn btn-success" href="{{ url_for('add_team') }}">➕ Добави</a>{% endif %}
</div>
<table class="table">
  <thead><tr><th>Име</th><th>Възраст</th><th>Пол</th><th>Играч(и)</th><th></th></tr></thead>
  <tbody>
    {% for t in teams %}
    <tr>
      <td>{{ t.name }}</td><td>{{ t.age_group or '-' }}</td><td>{{ t.gender or '-' }}</td><td>{{ t.players|length }}</td>
      <td>
        {% if current_user.role=='admin' %}
          <a class="btn btn-sm btn-primary" href="{{ url_for('edit_team', team_id=t.id) }}">✏</a>
          <form method="post" action="{{ url_for('delete_team', team_id=t.id) }}" style="display:inline" onsubmit="return confirm('Сигурни ли сте?')">
            <button class="btn btn-sm btn-danger">Изтрий</button>
          </form>
        {% endif %}
      </td>
    </tr>
    {% endfor %}
  </tbody>
</table>
{% endblock %}"""

BASIC_TEMPLATES['trainings.html'] = """{% extends 'base.html' %}{% block content %}
<div class="d-flex justify-content-between align-items-center mb-3">
  <h3>Тренировки</h3>
  <a class="btn btn-success" href="{{ url_for('add_training') }}">➕ Добави тренировка</a>
</div>
<table class="table">
  <thead><tr><th>Дата</th><th>Отбор</th><th>Час</th><th>Бележки</th><th>Присъствие</th></tr></thead>
  <tbody>
    {% for tr in trainings %}
    <tr>
      <td>{{ tr.date.strftime('%d.%m.%Y') }}</td>
      <td>{{ tr.session_team.name if tr.session_team else '—' }}</td>
      <td>{{ tr.start_time or '-' }} - {{ tr.end_time or '-' }}</td>
      <td>{{ tr.notes or '' }}</td>
      <td><a class="btn btn-sm btn-primary" href="{{ url_for('attendance_form', training_id=tr.id) }}">Отбележи</a></td>
    </tr>
    {% endfor %}
  </tbody>
</table>
{% endblock %}"""

BASIC_TEMPLATES['training_form.html'] = """{% extends 'base.html' %}{% block content %}
<h3>Добави/Редактирай тренировка</h3>
<form method="post" class="row g-2">
  <div class="col-6"><label>Дата</label><input name="date" type="date" class="form-control" required value="{{ training.date if training else '' }}"></div>
  <div class="col-6"><label>Отбор</label><select name="team_id" class="form-select">{% for t in teams %}<option value="{{ t.id }}" {% if training and training.team_id==t.id %}selected{% endif %}>{{ t.name }}</option>{% endfor %}</select></div>
  <div class="col-6"><label>Начален час</label><input name="start_time" class="form-control" placeholder="18:00" value="{{ training.start_time if training else '' }}"></div>
  <div class="col-6"><label>Краен час</label><input name="end_time" class="form-control" placeholder="19:30" value="{{ training.end_time if training else '' }}"></div>
  <div class="col-12"><label>Бележки</label><textarea name="notes" class="form-control">{{ training.notes if training else '' }}</textarea></div>
  <div class="col-12"><button class="btn btn-primary">Запази</button></div>
</form>
{% endblock %}"""

BASIC_TEMPLATES['attendance_form.html'] = """{% extends 'base.html' %}{% block content %}
<h3>Присъствие за {{ training.date.strftime('%d.%m.%Y') }} — {{ training.session_team.name if training.session_team else '' }}</h3>
<form method="post">
  <table class="table">
    <thead><tr><th>Играч</th><th>Присъства</th></tr></thead>
    <tbody>
      {% for p in players %}
      <tr>
        <td>{{ p.full_name }}</td>
        <td><input type="checkbox" name="present_{{ p.id }}" {% if attendance_map.get(p.id) == 'present' %}checked{% endif %}></td>
      </tr>
      {% endfor %}
    </tbody>
  </table>
  <div class="mb-3"><label>Бележка до родител (по избор)</label><input name="note" class="form-control"></div>
  <div class="mb-3">
    <button class="btn btn-primary">Запази присъствие</button>
    <button type="submit" name="notify" value="1" class="btn btn-warning">Запази и Изпрати уведомления</button>
  </div>
</form>
{% endblock %}"""

BASIC_TEMPLATES['attendance_stats.html'] = """{% extends 'base.html' %}{% block content %}
<h3>Статистика присъствия</h3>
<h5>Играч: {{ player.full_name }}</h5>
<p>Общо тренировки: {{ total }}, Присъствия: {{ present }}, Процент: {{ percent }}%</p>
<hr>
<h5>Отбор: {{ team.name if team else '-' }}</h5>
<table class="table">
  <thead><tr><th>Играч</th><th>Общо</th><th>Присъства</th><th>%</th></tr></thead>
  <tbody>
    {% for r in rows %}
    <tr><td>{{ r.full_name }}</td><td>{{ r.total }}</td><td>{{ r.present }}</td><td>{{ r.percent }}</td></tr>
    {% endfor %}
  </tbody>
</table>
{% endblock %}"""

BASIC_TEMPLATES['admin.html'] = """{% extends 'base.html' %}{% block content %}
<h3>Админ</h3>
<div class="mb-3"><a class="btn btn-primary" href="{{ url_for('create_user') }}">Добави потребител</a></div>
<div class="card mb-3"><div class="card-body">
<h5>Импорт на състезатели (.csv)</h5>
<form method="post" action="{{ url_for('admin_import') }}" enctype="multipart/form-data">
  <input type="file" name="file" accept=".csv" required class="form-control mb-2">
  <button class="btn btn-success">Качи и импортирай</button>
</form>
<small>Колони: Състезател,Отбор,Дата на раждане,Телефон на състезател,Телефон на родителя,Имейл</small>
</div></div>
<table class="table"><thead><tr><th>Потребител</th><th>Роля</th></tr></thead><tbody>{% for u in users %}<tr><td>{{ u.username }}</td><td>{{ u.role }}</td></tr>{% endfor %}</tbody></table>
{% endblock %}"""

# Write templates if missing
def ensure_templates():
    for name, content in BASIC_TEMPLATES.items():
        path = TEMPLATES_DIR / name
        if not path.exists():
            path.write_text(content, encoding='utf-8')
            logger.info(f'Created template {name}')

# -------------------- Helpers for notifications --------------------
def notify_attendance_change(player, training, status, note=None):
    # status: 'present' or 'absent'
    when = training.date.strftime('%d.%m.%Y')
    time_range = ''
    if training.start_time or training.end_time:
        time_range = f' ({training.start_time or ""} - {training.end_time or ""})'
    message = f'🏐 {player.full_name} е {"✅ присъствал(а)" if status=="present" else "❌ отсъствал(а)"} на тренировка на {when}{time_range}.'
    if note:
        message += f'\n📝 Бележка: {note}'
    # send only telegram for attendance notifications
    send_telegram(player.parent_telegram_id, message)

# -------------------- Routes --------------------
@app.route('/')
@login_required
def index():
    return render_template('index.html')

@app.route('/login', methods=['GET','POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password):
            login_user(user)
            flash('Успешен вход', 'success')
            return redirect(url_for('index'))
        flash('Грешно потребителско име или парола', 'error')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Изход', 'success')
    return redirect(url_for('login'))

# ---------- Teams ----------
@app.route('/teams')
@login_required
def teams():
    teams = Team.query.order_by(Team.name).all()
    return render_template('teams.html', teams=teams)

@app.route('/teams/add', methods=['GET','POST'])
@role_required('admin')
def add_team():
    if request.method == 'POST':
        name = request.form.get('name')
        age_group = request.form.get('age_group')
        gender = request.form.get('gender')
        if not name:
            flash('Името е задължително', 'error')
            return redirect(url_for('add_team'))
        t = Team(name=name, age_group=age_group, gender=gender)
        db.session.add(t); db.session.commit()
        flash('Отборът е добавен', 'success')
        return redirect(url_for('teams'))
    # simple quick form
    return '''<form method="post">Име: <input name="name"><br>Група: <input name="age_group"><br>Пол: <select name="gender"><option value="">--</option><option value="boys">boys</option><option value="girls">girls</option></select><button>Добави</button></form>'''

@app.route('/teams/<int:team_id>/edit', methods=['GET','POST'])
@role_required('admin')
def edit_team(team_id):
    t = Team.query.get_or_404(team_id)
    if request.method == 'POST':
        t.name = request.form.get('name'); t.age_group = request.form.get('age_group'); t.gender = request.form.get('gender')
        db.session.commit(); flash('Отборът е обновен', 'success'); return redirect(url_for('teams'))
    return f'''<form method="post">Име: <input name="name" value="{t.name}"><br>Група: <input name="age_group" value="{t.age_group or ''}"><br>Пол: <select name="gender"><option value="" {"selected" if not t.gender else ""}>--</option><option value="boys" {"selected" if t.gender=="boys" else ""}>boys</option><option value="girls" {"selected" if t.gender=="girls" else ""}>girls</option></select><button>Запази</button></form>'''

@app.route('/teams/<int:team_id>/delete', methods=['POST'])
@role_required('admin')
def delete_team(team_id):
    t = Team.query.get_or_404(team_id)
    for p in t.players:
        p.team_id = None
    db.session.delete(t); db.session.commit()
    flash('Отборът е изтрит', 'success')
    return redirect(url_for('teams'))
    
@app.route('/players/<int:player_id>/delete', methods=['POST'])
@role_required('trainer')
def delete_player(player_id):
    player = Player.query.get_or_404(player_id)

    # Изтриваме свързаните плащания и присъствия, за да няма осиротели записи
    Payment.query.filter_by(player_id=player.id).delete()
    Attendance.query.filter_by(player_id=player.id).delete()

    db.session.delete(player)
    db.session.commit()

    flash(f"Състезателят {player.full_name} е изтрит.", "success")
    return redirect(url_for('players'))


# ---------- Players ----------
@app.route('/players')
@login_required
def players():
    q = request.args.get('q', '').strip()
    team_id = request.args.get('team_id')
    query = Player.query
    if q:
        query = query.filter(Player.full_name.ilike(f'%{q}%'))
    if team_id and team_id.isdigit():
        query = query.filter_by(team_id=int(team_id))
    players = query.order_by(Player.full_name).all()
    return render_template('players.html', players=players)

@app.route('/players/add', methods=['GET','POST'])
@role_required('trainer')
def add_player():
    if request.method == 'POST':
        bd = request.form.get('birth_date')
        p = Player(
            full_name=request.form['full_name'],
            birth_date=None,
            player_phone=request.form.get('player_phone'),
            parent_phone=request.form.get('parent_phone'),
            parent_telegram_id=request.form.get('parent_telegram_id'),
            email=request.form.get('email'),
            team_id=request.form.get('team_id') or None,
            notes=request.form.get('notes')
        )
        if bd:
            try:
                p.birth_date = datetime.fromisoformat(bd).date()
            except Exception:
                p.birth_date = None
        db.session.add(p); db.session.commit()
        flash('Играчът е добавен', 'success'); return redirect(url_for('players'))
    teams = Team.query.order_by(Team.name).all()
    return render_template('player_form.html', player=None, teams=teams)

@app.route('/players/<int:player_id>/edit', methods=['GET','POST'])
@role_required('trainer')
def edit_player(player_id):
    player = Player.query.get_or_404(player_id)
    if request.method == 'POST':
        player.full_name = request.form['full_name']
        bd = request.form.get('birth_date')
        if bd:
            try:
                player.birth_date = datetime.fromisoformat(bd).date()
            except:
                player.birth_date = None
        player.player_phone = request.form.get('player_phone')
        player.parent_phone = request.form.get('parent_phone')
        player.parent_telegram_id = request.form.get('parent_telegram_id')
        player.email = request.form.get('email')
        player.team_id = request.form.get('team_id') or None
        player.notes = request.form.get('notes')
        db.session.commit(); flash('Играчът е обновен', 'success'); return redirect(url_for('players'))
    teams = Team.query.order_by(Team.name).all()
    return render_template('player_form.html', player=player, teams=teams)



# ---------- Payments ----------
@app.route('/payments')
@login_required
@role_required('trainer')
def payments_list():
    from datetime import datetime

    month = request.args.get('month', type=int) or datetime.now().month
    year = request.args.get('year', type=int) or datetime.now().year

    players = Player.query.order_by(Player.full_name).all()
    data = []
    new_payments = []

    print(f"[DEBUG] Намерени {len(players)} състезатели")

    for p in players:
        pay = Payment.query.filter_by(
            player_id=p.id,
            month=month,
            year=year
        ).first()

        # Ако няма плащане за този месец/година → създаваме го
        if not pay:
            pay = Payment(
                player_id=p.id,
                month=month,
                year=year,
                amount=0,
                status='pending'
            )
            db.session.add(pay)
            new_payments.append(pay)

        data.append({
            'player': p,
            'payment': pay,
            'overdue': (
                pay.status == 'pending'
                and datetime.now().day > 5
                and month == datetime.now().month
                and year == datetime.now().year
            )
        })

    # Записваме новите плащания, ако има такива
    if new_payments:
        db.session.commit()
        print(f"[DEBUG] Добавени нови плащания: {len(new_payments)}")

    return render_template(
        'payments.html',
        data=data,
        month=month,
        year=year
    )


@app.route('/payments/<int:payment_id>/mark_paid', methods=['POST'])
@login_required
@role_required('trainer')
def mark_payment_paid(payment_id):
    payment = Payment.query.get_or_404(payment_id)
    payment.mark_paid()
    db.session.commit()

    # известяване
    player = payment.player
    date_str = f"{payment.month:02d}.{payment.year}"
    msg = f"✅ Плащането за {date_str} е отбелязано като получено."

    if player.email:
        send_email(player.email, "Потвърждение за плащане", msg)
    if player.parent_telegram_id:
        send_telegram(player.parent_telegram_id, msg)

    flash(f"Плащането за {player.full_name} е маркирано като получено.", "success")
    return redirect(url_for('payments', year=payment.year, month=payment.month))


@app.route('/players/<int:player_id>/remind', methods=['GET'])
@login_required
def remind_player(player_id):
    player = Player.query.get_or_404(player_id)

    msg = f"Напомняне: Здравейте, родител на {player.full_name}, имате съобщение от треньора."
    
    # Изпращане по имейл
    send_email(player.email, "Напомняне от треньора", msg)

    # Изпращане по телеграм
    send_telegram(player.parent_telegram_id, msg)

    flash(f"Изпратено напомняне на {player.full_name}", "success")
    return redirect(url_for('players'))

@app.route('/attendance/<int:training_id>', methods=['GET', 'POST'])
@login_required
@role_required('trainer')
def attendance(training_id):
    training = TrainingSession.query.get_or_404(training_id)
    players = Player.query.filter_by(team_id=training.team_id).all()

    if request.method == 'POST':
        # изтриваме старите записи за това занимание
        Attendance.query.filter_by(session_id=training.id).delete()

        for player in players:
            status = 'present' if request.form.get(f'attendance_{player.id}') else 'absent'
            new_att = Attendance(session_id=training.id, player_id=player.id, status=status)
            db.session.add(new_att)
            db.session.commit()

            # Уведомяване на родителите
            date_str = training.date.strftime('%d.%m.%Y')
            if status == 'present':
                msg = f"✅ {player.full_name} присъства на тренировка на {date_str}."
            else:
                msg = f"❌ {player.full_name} отсъства от тренировка на {date_str}."

            # Email
            if player.email:
                send_email(player.email, "Известие за присъствие", msg)

            # Telegram
            if player.parent_telegram_id:
                send_telegram(player.parent_telegram_id, msg)

        flash("Присъствията са записани и родителите са уведомени.", "success")
        return redirect(url_for('trainings'))

    # Съществуващи присъствия за предварително маркиране
    attendance_map = {
        a.player_id: a.status
        for a in Attendance.query.filter_by(session_id=training.id).all()
    }

    return render_template('attendance.html', training=training, players=players, attendance=attendance_map)


@app.route('/payments/<int:payment_id>/remind')
@login_required
@role_required('trainer')
def remind_payment(payment_id):
    payment = Payment.query.get_or_404(payment_id)
    player = payment.player

    # Съобщение за родителите
    message = f"Напомняне: Таксата за {payment.month}/{payment.year} за {player.full_name} е неплатена."

    # Изпращане на имейл
    if player.email:
        send_email(player.email, "Напомняне за плащане", message)

    # Изпращане на телеграм
    if player.parent_telegram_id:
        send_telegram(player.parent_telegram_id, message)

    flash(f'Изпратено е напомняне за {player.full_name} 🔔', 'success')
    return redirect(url_for('payments_list'))

@app.route('/payments/remind_all')
@login_required
@role_required('trainer')
def remind_all_payments():
    payments = Payment.query.all()
    for payment in payments:
        player = payment.player
        message = f"Напомняне: Таксата за {payment.month}/{payment.year} за {player.full_name}."
        
        if player.email:
            send_email(player.email, "Напомняне за плащане", message)
        if player.parent_telegram_id:
            send_telegram(player.parent_telegram_id, message)

    flash("Изпратени са напомняния на всички 📨", "success")
    return redirect(url_for('payments_list'))

@app.route('/stats')
@login_required
@role_required('trainer')
def stats_page():
    # --- Плащания ---
    payments_stats = db.session.query(
        Payment.month,
        Payment.year,
        db.func.count(Payment.id).label("total"),
        db.func.sum(case((Payment.status == 'paid', 1), else_=0)).label("paid"),
        db.func.sum(case((Payment.status != 'paid', 1), else_=0)).label("unpaid")
    ).group_by(Payment.year, Payment.month).order_by(Payment.year, Payment.month).all()

    payments_labels = [f"{p.month:02d}/{p.year}" for p in payments_stats]
    payments_paid = [p.paid for p in payments_stats]
    payments_unpaid = [p.unpaid for p in payments_stats]

    # --- Присъствия ---
    attendance_stats = db.session.query(
        extract('month', TrainingSession.date).label("month"),
        extract('year', TrainingSession.date).label("year"),
        db.func.avg(case((Attendance.status == 'present', 1), else_=0)).label("attendance_percent")
    ).join(Attendance, Attendance.session_id == TrainingSession.id) \
     .group_by("year", "month") \
     .order_by("year", "month").all()

    attendance_labels = [f"{int(row.month):02d}/{int(row.year)}" for row in attendance_stats]
    attendance_percent = [round(row.attendance_percent * 100, 1) if row.attendance_percent else 0 for row in attendance_stats]

    return render_template(
        "stats.html",
        payments_labels=payments_labels,
        payments_paid=payments_paid,
        payments_unpaid=payments_unpaid,
        attendance_labels=attendance_labels,
        attendance_percent=attendance_percent
    )

import csv
from io import StringIO
from flask import Response

@app.route('/stats/payments_csv')
@login_required
@role_required('trainer')
def stats_payments_csv():
    # Взимаме всички уникални години и месеци, сортирани
    periods = db.session.query(
        Payment.year, Payment.month
    ).distinct().order_by(Payment.year, Payment.month).all()

    # Взимаме всички състезатели
    players = Player.query.order_by(Player.full_name).all()

    # Подготвяме данни: {player_id: { (year, month): amount }}
    payments_map = {}
    for p in players:
        payments_map[p.id] = {}
        for year, month in periods:
            pay = Payment.query.filter_by(player_id=p.id, year=year, month=month).first()
            payments_map[p.id][(year, month)] = float(pay.amount) if (pay and pay.status == 'paid') else 0.0

    # Генерираме CSV
    output = StringIO()
    writer = csv.writer(output)

    # Заглавен ред
    header = ["Състезател"] + [f"{month:02d}.{year}" for year, month in periods]
    writer.writerow(header)

    # Редове за състезатели
    for p in players:
        row = [p.full_name]
        for year, month in periods:
            row.append(payments_map[p.id][(year, month)])
        writer.writerow(row)

    # Връщаме като отговор за сваляне
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=payments_stats.csv"}
    )



# ---------- Admin ----------
@app.route('/admin')
@role_required('admin')
def admin_panel():
    users = User.query.order_by(User.username).all()
    return render_template('admin.html', users=users)

@app.route('/admin/users/create', methods=['GET','POST'])
@role_required('admin')
def create_user():
    if request.method == 'POST':
        username = request.form['username']; password = request.form['password']; role = request.form.get('role','trainer')
        if User.query.filter_by(username=username).first():
            flash('Потребител вече съществува','error'); return redirect(url_for('create_user'))
        u = User(username=username, role=role); u.set_password(password); db.session.add(u); db.session.commit()
        flash('Потребител създаден','success'); return redirect(url_for('admin_panel'))
    return render_template('create_user.html')

# ---------- Coaches (admin) ----------
@app.route('/admin/coaches', methods=['GET','POST'])
@role_required('admin')
def admin_coaches():
    if request.method == 'POST':
        # Създаване на нов треньор (User + CoachProfile)
        username = request.form.get('username')
        password = request.form.get('password')
        full_name = request.form.get('full_name')
        phone = request.form.get('phone')
        team_ids = request.form.getlist('team_ids')
        if not username or not password:
            flash('Потребителско име и парола са задължителни','error')
            return redirect(url_for('admin_coaches'))
        if User.query.filter_by(username=username).first():
            flash('Потребител вече съществува','error')
            return redirect(url_for('admin_coaches'))
        u = User(username=username, role='trainer')
        u.set_password(password)
        db.session.add(u); db.session.commit()
        cp = CoachProfile(user_id=u.id, full_name=full_name, phone=phone,
                          can_manage_players=bool(request.form.get('can_manage_players')),
                          can_manage_payments=bool(request.form.get('can_manage_payments')),
                          can_mark_attendance=bool(request.form.get('can_mark_attendance')),
                          can_manage_slots=bool(request.form.get('can_manage_slots')),
                          can_manage_tournaments=bool(request.form.get('can_manage_tournaments')),
                          can_manage_inventory=bool(request.form.get('can_manage_inventory')))
        db.session.add(cp); db.session.commit()
        for tid in team_ids:
            if tid.isdigit():
                db.session.add(CoachTeam(coach_id=cp.id, team_id=int(tid)))
        db.session.commit()
        flash('Треньорът е създаден','success')
        return redirect(url_for('admin_coaches'))
    coaches = CoachProfile.query.all()
    coach_rows = []
    for c in coaches:
        user = User.query.get(c.user_id)
        teams = [Team.query.get(ct.team_id).name for ct in CoachTeam.query.filter_by(coach_id=c.id).all() if Team.query.get(ct.team_id)]
        coach_rows.append({'coach': c, 'user': user, 'teams': teams})
    teams = Team.query.order_by(Team.name).all()
    return render_template('coaches.html', coach_rows=coach_rows, teams=teams)

@app.route('/admin/coaches/<int:coach_id>/teams', methods=['POST'])
@role_required('admin')
def admin_coach_set_teams(coach_id):
    CoachTeam.query.filter_by(coach_id=coach_id).delete()
    for tid in request.form.getlist('team_ids'):
        if tid.isdigit():
            db.session.add(CoachTeam(coach_id=coach_id, team_id=int(tid)))
    db.session.commit()
    flash('Отборите на треньора са обновени','success')
    return redirect(url_for('admin_coaches'))

@app.route('/admin/seed_coaches', methods=['POST'])
@role_required('admin')
def seed_default_coaches():
    # Създава профили за Anatoli и Pepi (ако липсват)
    def ensure(username, password, full_name):
        u = User.query.filter_by(username=username).first()
        if not u:
            u = User(username=username, role='trainer')
            u.set_password(password)
            db.session.add(u); db.session.commit()
        cp = CoachProfile.query.filter_by(user_id=u.id).first()
        if not cp:
            cp = CoachProfile(user_id=u.id, full_name=full_name, can_manage_players=True,
                              can_manage_payments=True, can_mark_attendance=True,
                              can_manage_slots=False, can_manage_tournaments=True,
                              can_manage_inventory=True)
            db.session.add(cp); db.session.commit()
    ensure('anatoli', 'anatoli9010', 'Anatoli')
    ensure('pepi', 'pepi2025', 'Pepi')
    flash('Треньорите Anatoli и Pepi са налични/обновени','success')
    return redirect(url_for('admin_coaches'))

@app.route('/coaches/<int:coach_id>/schedule')
@login_required
def coach_schedule(coach_id):
    # график за треньор: всички тренировки на отборите, към които е зачислен
    month = request.args.get('month', type=int) or date.today().month
    year = request.args.get('year', type=int) or date.today().year
    team_ids = [ct.team_id for ct in CoachTeam.query.filter_by(coach_id=coach_id).all()]
    q = TrainingSession.query 
    q = q.filter(extract('month', TrainingSession.date)==month,
                 extract('year', TrainingSession.date)==year)
    if team_ids:
        q = q.filter(TrainingSession.team_id.in_(team_ids))
    trainings = q.order_by(TrainingSession.date.asc(), TrainingSession.start_time.asc()).all()
    for t in trainings:
        t.session_team = Team.query.get(t.team_id) if t.team_id else None
    return render_template('trainings.html', trainings=trainings, month=month, year=year, teams=Team.query.order_by(Team.name).all(), selected_team_id=None)

# ---------- Tournament management ----------
class Tournament(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    start_date = db.Column(db.Date, nullable=True)
    end_date = db.Column(db.Date, nullable=True)
    venue = db.Column(db.String(120), nullable=True)
    notes = db.Column(db.String(400), nullable=True)

class TournamentTeam(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    tournament_id = db.Column(db.Integer, db.ForeignKey('tournament.id'), nullable=False)
    team_id = db.Column(db.Integer, db.ForeignKey('team.id'), nullable=False)

@app.route('/tournaments', methods=['GET','POST'])
@login_required
@coach_permission_required('can_manage_tournaments')
def tournaments():
    if request.method == 'POST':
        t = Tournament(
            name=request.form.get('name'),
            start_date=datetime.fromisoformat(request.form.get('start_date')).date() if request.form.get('start_date') else None,
            end_date=datetime.fromisoformat(request.form.get('end_date')).date() if request.form.get('end_date') else None,
            venue=request.form.get('venue'),
            notes=request.form.get('notes')
        )
        db.session.add(t); db.session.commit()
        for tid in request.form.getlist('team_ids'):
            if tid.isdigit():
                db.session.add(TournamentTeam(tournament_id=t.id, team_id=int(tid)))
        db.session.commit()
        flash('Турнирът е добавен','success')
        return redirect(url_for('tournaments'))
    items = Tournament.query.order_by(Tournament.start_date.desc().nullslast()).all()
    teams = Team.query.order_by(Team.name).all()
    rows = []
    for t in items:
        t_teams = [Team.query.get(x.team_id).name for x in TournamentTeam.query.filter_by(tournament_id=t.id).all() if Team.query.get(x.team_id)]
        rows.append({'t': t, 'teams': t_teams})
    return render_template('tournaments.html', tournaments=rows, teams=teams)

@app.route('/tournaments/<int:t_id>/delete', methods=['POST'])
@login_required
@coach_permission_required('can_manage_tournaments')
def delete_tournament(t_id):
    TournamentTeam.query.filter_by(tournament_id=t_id).delete()
    Tournament.query.filter_by(id=t_id).delete()
    db.session.commit()
    flash('Турнирът е изтрит','success')
    return redirect(url_for('tournaments'))

# ---------- Inventory management ----------
class InventoryItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    quantity = db.Column(db.Integer, nullable=False, default=0)
    location = db.Column(db.String(120), nullable=True)  # НУПИ/Чавдар/Стадион/склад
    condition = db.Column(db.String(120), nullable=True) # добро/за ремонт/износено
    notes = db.Column(db.String(400), nullable=True)

@app.route('/inventory', methods=['GET','POST'])
@login_required
@coach_permission_required('can_manage_inventory')
def inventory():
    if request.method == 'POST':
        it = InventoryItem(
            name=request.form.get('name'),
            quantity=request.form.get('quantity', type=int) or 0,
            location=request.form.get('location'),
            condition=request.form.get('condition'),
            notes=request.form.get('notes')
        )
        db.session.add(it); db.session.commit()
        flash('Артикулът е добавен','success')
        return redirect(url_for('inventory'))
    items = InventoryItem.query.order_by(InventoryItem.name).all()
    return render_template('inventory.html', items=items)

@app.route('/inventory/<int:item_id>/delete', methods=['POST'])
@login_required
@coach_permission_required('can_manage_inventory')
def inventory_delete(item_id):
    InventoryItem.query.filter_by(id=item_id).delete(); db.session.commit()
    flash('Артикулът е изтрит','success')
    return redirect(url_for('inventory'))

@app.route('/inventory/<int:item_id>/adjust', methods=['POST'])
@login_required
@coach_permission_required('can_manage_inventory')
def inventory_adjust(item_id):
    item = InventoryItem.query.get_or_404(item_id)
    delta = request.form.get('delta', type=int)
    item.quantity = max(0, (item.quantity or 0) + (delta or 0))
    db.session.commit()
    return redirect(url_for('inventory'))

# CSV import
def allowed_file(filename):
    ext = filename.rsplit('.',1)[-1].lower()
    return ext in ALLOWED_EXT

@app.route('/admin/import', methods=['POST'])
@role_required('admin')
def admin_import():
    if 'file' not in request.files:
        flash('Файлът не е качен','error'); return redirect(url_for('admin_panel'))
    f = request.files['file']
    if f.filename == '':
        flash('Няма избран файл','error'); return redirect(url_for('admin_panel'))
    if not allowed_file(f.filename):
        flash('Неразрешен тип (само CSV)','error'); return redirect(url_for('admin_panel'))
    filename = secure_filename(f.filename); path = os.path.join(app.config['UPLOAD_FOLDER'], filename); f.save(path)
    try:
        df = pd.read_csv(path)
    except Exception:
        flash('Грешка при четене','error'); return redirect(url_for('admin_panel'))
    col_map = {c.strip():c for c in df.columns}
    required = ['Състезател','Отбор','Дата на раждане','Телефон на състезател','Телефон на родителя','Имейл']
    missing = [r for r in required if r not in col_map]
    if missing:
        flash(f'Липсват колони: {", ".join(missing)}','error'); return redirect(url_for('admin_panel'))
    created=0; updated=0
    for _,row in df.iterrows():
        full_name = str(row[col_map['Състезател']]).strip()
        team_name = str(row[col_map['Отбор']]).strip()
        dob_raw = row[col_map['Дата на раждане']]
        player_phone = str(row[col_map['Телефон на състезател']]).strip()
        parent_phone = str(row[col_map['Телефон на родителя']]).strip()
        email = str(row[col_map['Имейл']]).strip()
        dob=None
        if pd.notna(dob_raw):
            try:
                if isinstance(dob_raw,str): dob = datetime.fromisoformat(dob_raw).date()
                else: dob = pd.to_datetime(dob_raw).date()
            except: dob=None
        team=None
        if team_name:
            team = Team.query.filter_by(name=team_name).first()
            if not team:
                age_group=None; gender=None; low = team_name.lower()
                for g in ['u12','u13','u14','u16','u18']:
                    if g in low: age_group=g
                if 'girl' in low or 'жен' in low or 'момич' in low: gender='girls'
                if 'boy' in low or 'момч' in low or 'мъж' in low: gender='boys'
                team = Team(name=team_name, age_group=age_group, gender=gender); db.session.add(team); db.session.commit()
        existing=None
        if parent_phone and parent_phone!='nan':
            existing = Player.query.filter_by(full_name=full_name, parent_phone=parent_phone).first()
        if not existing and player_phone and player_phone!='nan':
            existing = Player.query.filter_by(full_name=full_name, player_phone=player_phone).first()
        if existing:
            existing.birth_date = dob or existing.birth_date
            existing.player_phone = player_phone or existing.player_phone
            existing.parent_phone = parent_phone or existing.parent_phone
            existing.email = email or existing.email
            existing.team_id = team.id if team else existing.team_id
            db.session.commit(); updated+=1
        else:
            newp = Player(full_name=full_name, birth_date=dob, player_phone=player_phone, parent_phone=parent_phone, email=email, team_id=team.id if team else None)
            db.session.add(newp); db.session.commit(); created+=1
    flash(f'Импорт: добавени {created}, обновени {updated}','success'); return redirect(url_for('admin_panel'))

# ---------- Trainings & Attendance ----------
@app.route('/trainings')
@login_required
def trainings():
    # Филтри по месец/година и отбор
    today = date.today()
    month = request.args.get('month', type=int) or today.month
    year = request.args.get('year', type=int) or today.year
    team_id = request.args.get('team_id', type=int)

    q = TrainingSession.query
    # по месец/година
    q = q.filter(
        extract('month', TrainingSession.date) == month,
        extract('year', TrainingSession.date) == year
    )
    if team_id:
        q = q.filter(TrainingSession.team_id == team_id)

    trainings = q.order_by(TrainingSession.date.asc(), TrainingSession.start_time.asc()).all()
    # attach team for display convenience
    for t in trainings:
        t.session_team = Team.query.get(t.team_id) if t.team_id else None

    teams = Team.query.order_by(Team.name).all()
    return render_template('trainings.html', trainings=trainings, month=month, year=year, teams=teams, selected_team_id=team_id)

@app.route('/trainings/add', methods=['GET', 'POST'])
@login_required
def add_training():
    if request.method == 'POST':
        try:
            d = request.form.get('date')
            team_id_raw = request.form.get('team_id')
            if not team_id_raw:
                flash('Моля, изберете отбор', 'error')
                return redirect(url_for('add_training'))

            team_id = int(team_id_raw)
            start_time = request.form.get('start_time') or ''
            end_time = request.form.get('end_time') or ''
            notes = request.form.get('notes') or ''

            tr = TrainingSession(
                team_id=team_id,
                date=datetime.fromisoformat(d).date(),
                start_time=start_time,
                end_time=end_time,
                notes=notes
            )
            db.session.add(tr)
            db.session.commit()

            flash('Тренировка добавена', 'success')
            return redirect(url_for('trainings'))

        except Exception as e:
            logger.exception('Add training failed')
            flash('Грешка при добавяне на тренировка: ' + str(e), 'error')
            return redirect(url_for('add_training'))

    teams = Team.query.order_by(Team.name).all()
    return render_template('add_training.html', teams=teams)


@app.route('/trainings/<int:training_id>/edit', methods=['GET','POST'])
@role_required('trainer')
def edit_training(training_id):
    tr = TrainingSession.query.get_or_404(training_id)
    if request.method == 'POST':
        tr.date = datetime.fromisoformat(request.form.get('date')).date()
        tr.team_id = int(request.form.get('team_id'))
        tr.start_time = request.form.get('start_time')
        tr.end_time = request.form.get('end_time')
        tr.notes = request.form.get('notes')
        db.session.commit(); flash('Тренировката е обновена','success'); return redirect(url_for('trainings'))
    teams = Team.query.order_by(Team.name).all()
    return render_template('training_form.html', teams=teams, training=tr)

@app.route('/trainings/<int:training_id>/delete', methods=['POST'])
@role_required('trainer')
def delete_training(training_id):
    tr = TrainingSession.query.get_or_404(training_id)
    db.session.delete(tr); db.session.commit(); flash('Тренировката е изтрита','success'); return redirect(url_for('trainings'))

@app.route('/trainings/<int:training_id>/attendance', methods=['GET','POST'])
@role_required('trainer')
def attendance_form(training_id):
    training = TrainingSession.query.get_or_404(training_id)
    players = Player.query.filter_by(team_id=training.team_id).order_by(Player.full_name).all()
    # map existing attendance
    existing = Attendance.query.filter_by(session_id=training.id).all()
    attendance_map = {a.player_id: a.status for a in existing}
    if request.method == 'POST':
        note = request.form.get('note','').strip()
        notify = request.form.get('notify') == '1'
        # update/create attendance rows
        changed = []
        for p in players:
            present_key = f'present_{p.id}'
            is_present = present_key in request.form
            new_status = 'present' if is_present else 'absent'
            a = Attendance.query.filter_by(session_id=training.id, player_id=p.id).first()
            if a:
                if a.status != new_status:
                    a.status = new_status
                    a.noted_at = datetime.now(timezone.utc)
                    db.session.commit()
                    changed.append((p, new_status))
            else:
                a = Attendance(session_id=training.id, player_id=p.id, status=new_status, noted_at=datetime.now(timezone.utc))
                db.session.add(a); db.session.commit()
                changed.append((p, new_status))
        # if notify requested, send messages
        if notify:
            for p, st in changed:
                try:
                    notify_attendance_change(p, training, st, note or None)
                except Exception:
                    logger.exception('Notify attendance failed for %s', p.id)
        flash('Присъствия записани' + (', уведомления изпратени' if notify else ''), 'success')
        return redirect(url_for('trainings'))
    # add training_team for template
    training.session_team = Team.query.get(training.team_id) if training.team_id else None
    return render_template('attendance_form.html', training=training, players=players, attendance_map=attendance_map)

# Attendance statistics
@app.route('/attendance/stats/player/<int:player_id>')
@login_required
def attendance_stats_player(player_id):
    player = Player.query.get_or_404(player_id)
    total = Attendance.query.filter_by(player_id=player_id).count()
    present = Attendance.query.filter_by(player_id=player_id, status='present').count()
    percent = round((present / total * 100) if total>0 else 0,1)
    return render_template('attendance_stats.html', player=player, total=total, present=present, percent=percent, team=None, rows=[])

@app.route('/attendance/stats/team/<int:team_id>')
@login_required
def attendance_stats_team(team_id):
    team = Team.query.get_or_404(team_id)
    players = Player.query.filter_by(team_id=team_id).order_by(Player.full_name).all()
    rows = []
    for p in players:
        total = Attendance.query.filter_by(player_id=p.id).count()
        present = Attendance.query.filter_by(player_id=p.id, status='present').count()
        percent = round((present / total * 100) if total>0 else 0,1)
        rows.append({'full_name': p.full_name, 'total': total, 'present': present, 'percent': percent})
    return render_template('attendance_stats.html', player=None, total=0, present=0, percent=0, team=team, rows=rows)

import csv
from io import StringIO
from flask import Response

@app.route('/stats/export/payments', methods=['GET', 'POST'])
@login_required
@role_required('trainer')
def export_stats_payments():
    # Четем параметри от URL: ?start=2025-01&end=2025-03
    start_str = request.args.get('start')
    end_str = request.args.get('end')

    # Парсваме периода
    if start_str and end_str:
        start_year, start_month = map(int, start_str.split('-'))
        end_year, end_month = map(int, end_str.split('-'))
    else:
        # Ако няма подаден период, взимаме целия
        first_payment = Payment.query.order_by(Payment.year, Payment.month).first()
        last_payment = Payment.query.order_by(Payment.year.desc(), Payment.month.desc()).first()
        start_year, start_month = first_payment.year, first_payment.month
        end_year, end_month = last_payment.year, last_payment.month

    # Всички играчи
    players = Player.query.order_by(Player.full_name).all()

    # Всички уникални (година, месец) в периода
    periods = db.session.query(Payment.year, Payment.month) \
        .filter(
            (Payment.year > start_year) | ((Payment.year == start_year) & (Payment.month >= start_month)),
            (Payment.year < end_year) | ((Payment.year == end_year) & (Payment.month <= end_month))
        ) \
        .group_by(Payment.year, Payment.month) \
        .order_by(Payment.year, Payment.month).all()

    # Речник за суми по (player_id, year, month)
    payments = Payment.query.filter(
        (Payment.year > start_year) | ((Payment.year == start_year) & (Payment.month >= start_month)),
        (Payment.year < end_year) | ((Payment.year == end_year) & (Payment.month <= end_month))
    ).all()

    payments_map = {
        (p.player_id, p.year, p.month): p.amount if p.status == 'paid' else 0
        for p in payments
    }

    # CSV файл
    si = StringIO()
    cw = csv.writer(si)
    header = ["Състезател"] + [f"{m:02d}.{y}" for y, m in periods]
    cw.writerow(header)

    for player in players:
        row = [player.full_name]
        for y, m in periods:
            row.append(payments_map.get((player.id, y, m), 0))
        cw.writerow(row)

    return Response(
        si.getvalue(),
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=payments_detailed.csv"}
    )

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from flask import send_file
import io

@app.route('/stats/export/payments_excel')
@login_required
@role_required('trainer')
def export_stats_payments_excel():
    # Извличаме всички плащания
    payments = db.session.query(
        Payment,
        Player.name.label('player_name')
    ).join(Player, Payment.player_id == Player.id).order_by(Payment.year, Payment.month).all()

    # Подготвяме структура {играч: {месец/година: сума}}
    data = {}
    months_set = set()

    for p, player_name in payments:
        key = f"{p.month:02d}/{p.year}"
        months_set.add(key)
        if player_name not in data:
            data[player_name] = {}
        data[player_name][key] = p.amount if p.status == 'paid' else 0

    months_list = sorted(months_set, key=lambda x: (int(x.split('/')[1]), int(x.split('/')[0])))

    # Създаваме Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Плащания"

    # Заглавен ред
    headers = ["Състезател"] + months_list
    ws.append(headers)

    # Данни
    for player_name, months in data.items():
        row = [player_name]
        for m in months_list:
            row.append(months.get(m, 0))
        ws.append(row)

    # Авто-ширина
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Запис в паметта
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True,
                     download_name="payments_stats.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route('/stats/export/payments/excel_v2')
@login_required
@role_required('trainer')
def export_stats_payments_excel_v2():
    start = request.args.get('start')
    end = request.args.get('end')

    query = db.session.query(Payment).join(Player)

    if start:
        start_year, start_month = map(int, start.split('-'))
        query = query.filter(
            (Payment.year > start_year) |
            ((Payment.year == start_year) & (Payment.month >= start_month))
        )

    if end:
        end_year, end_month = map(int, end.split('-'))
        query = query.filter(
            (Payment.year < end_year) |
            ((Payment.year == end_year) & (Payment.month <= end_month))
        )

    payments = query.order_by(Player.full_name, Payment.year, Payment.month).all()

    # Подготовка на структура {играч: {месец: сума}}
    from collections import defaultdict
    data = defaultdict(lambda: defaultdict(int))

    months_set = set()

    for p in payments:
        month_label = f"{p.month:02d}/{p.year}"
        months_set.add(month_label)
        data[p.player.full_name][month_label] = p.amount if p.status == 'paid' else 0

    months_list = sorted(months_set, key=lambda x: (int(x.split('/')[1]), int(x.split('/')[0])))

    # Създаване на Excel
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Плащания"

    # Заглавен ред
    ws.append(["Състезател"] + months_list)

    # Редове с данни
    for player_name in sorted(data.keys()):
        row = [player_name] + [data[player_name].get(m, 0) for m in months_list]
        ws.append(row)

    # Запис във временно място
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=payments_stats.xlsx"}
    )


@app.route('/stats/export/attendance')
@login_required
@role_required('trainer')
def export_stats_attendance():
    attendance_stats = db.session.query(
        extract('month', TrainingSession.date).label("month"),
        extract('year', TrainingSession.date).label("year"),
        db.func.avg(case((Attendance.status == 'present', 1), else_=0)).label("attendance_percent")
    ).join(Attendance, Attendance.session_id == TrainingSession.id) \
     .group_by("year", "month") \
     .order_by("year", "month").all()

    si = StringIO()
    cw = csv.writer(si)
    cw.writerow(["Месец", "Година", "Средно присъствие %"])
    for row in attendance_stats:
        cw.writerow([int(row.month), int(row.year), round(row.attendance_percent * 100, 1)])

    return Response(
        si.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=attendance_stats.csv"}
    )


# API bind for telegram bot
@app.route('/api/bind', methods=['POST'])


def normalize_phone(num: str) -> str:
    if not num:
        return ""
    num = re.sub(r'\D', '', num)  # маха всичко, което не е цифра
    if num.startswith("0") and len(num) == 10:  # 087..., 088...
        num = "359" + num[1:]
    elif num.startswith("00"):  # 00359...
        num = num[2:]
    return num

@app.route('/api/bind', methods=['POST'])
def api_bind():
    data = request.get_json(force=True, silent=True)
    if not data:
        return jsonify({"ok": False, "error": "no json"}), 400
    
    phone = data.get('phone')
    telegram_id = data.get('telegram_id')

    if not phone or not telegram_id:
        return jsonify({"ok": False, "error": "missing phone or telegram_id"}), 400

    normalized = normalize_phone(phone)
    phone_to_telegram[normalized] = str(telegram_id)

    # Търсене с нормализиран телефон в базата
    players = [
        p for p in Player.query.all()
        if normalize_phone(p.parent_phone) == normalized
    ]

    count = 0
    for p in players:
        p.parent_telegram_id = str(telegram_id)
        count += 1

    db.session.commit()
    logger.info(f'Bind: {normalized} -> {telegram_id}, matched {count}')
    return jsonify({"ok": True, "matched": count})

# -------------------- Scheduler --------------------
scheduler = BackgroundScheduler()

def send_monthly_reminders():
    today = date.today()
    month = today.month; year = today.year
    pending = Payment.query.filter_by(year=year, month=month, status='pending').all()
    logger.info(f'Pending {len(pending)} for {month}/{year}')
    for p in pending:
        try:
            send_email(p.player.email, 'Напомняне за плащане', f'Напомняне: плащане за {p.player.full_name} за {p.month}/{p.year}.')
            send_telegram(p.player.parent_telegram_id, f'Напомняне: плащане за {p.player.full_name} за {p.month}/{p.year}.')
        except:
            logger.exception('Failed monthly reminder')

# -------------------- Telegram polling bot --------------------
def telegram_polling_loop():
    if not TELEGRAM_BOT_TOKEN:
        logger.info("No Telegram token — skipping bot.")
        return
    logger.info("Starting Telegram polling bot")
    offset = None
    base = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}"
    phone_re = re.compile(r'(\+?\d[\d\s\-()]{4,}\d)')
    while True:
        try:
            params = {'timeout':20}
            if offset: params['offset'] = offset
            resp = requests.get(base + "/getUpdates", params=params, timeout=30)
            data = resp.json()
            if not data.get('ok'):
                time.sleep(2); continue
            for upd in data.get('result', []):
                offset = upd['update_id'] + 1
                msg = upd.get('message') or upd.get('edited_message')
                if not msg: continue
                chat = msg.get('chat', {}); chat_id = chat.get('id'); text = msg.get('text','').strip()
                if text.lower().startswith('/start'):
                    requests.post(base + "/sendMessage", json={"chat_id": chat_id, "text": "Здравейте! Изпратете номера на родителя (напр. +359888111222) за да се свържете."})
                    continue
                m = phone_re.search(text)
                if m:
                    phone = re.sub(r'[\s\-\(\)]','', m.group(1))
                    try:
                        j = {"phone": phone, "telegram_id": chat_id}
                        r = requests.post("http://127.0.0.1:5000/api/bind", json=j, timeout=5)
                        if r.ok and r.json().get('ok'):
                            matched = r.json().get('matched',0)
                            reply = f"Телефон {phone} е свързан успешно. Намерени играчи: {matched}."
                        else:
                            reply = f"Опит за свързване направен, но не е намерен състезател."
                    except Exception:
                        logger.exception('Local bind failed'); reply = "Грешка при свързване."
                    requests.post(base + "/sendMessage", json={"chat_id": chat_id, "text": reply})
                else:
                    requests.post(base + "/sendMessage", json={"chat_id": chat_id, "text": "Моля изпратете телефонния номер на родителя (напр. +359888111222)."})
        except Exception:
            logger.exception('Telegram polling error'); time.sleep(5)

bot_thread = None

def start_telegram_bot_thread():
    logger.info("Telegram bot thread not started — running in webhook mode.")


# -------------------- Sample data & init --------------------
def create_sample_data():
    if User.query.count() == 0:
        admin = User(username='admin', role='admin'); admin.set_password('admin')
        trener = User(username='trener', role='trainer'); trener.set_password('trainer')
        db.session.add_all([admin, trener]); db.session.commit(); logger.info('Sample users created')
    # create common teams if missing
    existing = Team.query.count()
    if existing == 0:
        groups = ['u12','u14','u16','u18']
        for g in groups:
            Team(name=f'U{g[1:]} Boys', age_group=g, gender='boys')
            Team(name=f'U{g[1:]} Girls', age_group=g, gender='girls')
        # commit
        db.session.commit()
    if Player.query.count() == 0:
        # pick some teams
        tb = Team.query.filter_by(name='U12 Boys').first()
        tg = Team.query.filter_by(name='U14 Girls').first()
        p1 = Player(full_name='Иван Иванов', birth_date=date(2012,5,4), player_phone='+359888100001', parent_phone='+359888111001', email='ivan.parent@example.com', team_id=tb.id if tb else None)
        p2 = Player(full_name='Георги Георгиев', birth_date=date(2011,7,12), player_phone='+359888100002', parent_phone='+359888111002', email='georgi.parent@example.com', team_id=tb.id if tb else None)
        p3 = Player(full_name='Мария Петрова', birth_date=date(2010,8,19), player_phone='+359888100003', parent_phone='+359888111003', email='maria.parent@example.com', team_id=tg.id if tg else None)
        db.session.add_all([p1,p2,p3]); db.session.commit()
        # add some payments
        for p in [p1,p2,p3]:
            pay = Payment(player_id=p.id, year=date.today().year, month=date.today().month, amount=30.0, status='pending')
            db.session.add(pay)
        # add a training
        if tb:
            tr = TrainingSession(team_id=tb.id, date=date.today(), start_time='18:00', end_time='19:30', notes='Първа тренировка')
            db.session.add(tr)
        db.session.commit()
        logger.info('Sample players, payments, trainings created')

def init_app():
    ensure_templates()
    with app.app_context():
        db.create_all()
        create_sample_data()
    scheduler.add_job(send_monthly_reminders, 'interval', minutes=60, id='monthly_reminders', replace_existing=True)
    scheduler.start()
    start_telegram_bot_thread()
    logger.info('App initialized')

@app.route('/payments/add', methods=['GET', 'POST'])
@login_required
def add_payment():
    if request.method == 'POST':
        player_id = int(request.form.get('player_id'))
        month = int(request.form.get('month'))
        year = int(request.form.get('year'))
        amount = float(request.form.get('amount', 0))

        # Създаваме плащането
        payment = Payment(
            player_id=player_id,
            month=month,
            year=year,
            amount=amount,
            status='pending'  # вместо is_paid
        )
        db.session.add(payment)
        db.session.commit()

        # Известяване
        player = Player.query.get(player_id)
        date_str = f"{month:02d}.{year}"
        message = f"💳 Добавено е ново плащане за {player.full_name} за {date_str} — {amount:.2f} лв."

        if player.email:
            send_email(player.email, "Ново плащане", message)
        if player.parent_telegram_id:
            send_telegram(player.parent_telegram_id, message)

        flash("✅ Плащането е добавено успешно и известието е изпратено", "success")
        return redirect(url_for('payments_list'))

    players = Player.query.all()
    return render_template('add_payment.html', players=players)



@app.route('/api/player/<int:player_id>/payments')
@login_required
def get_player_payments(player_id):
    payments = Payment.query.filter_by(player_id=player_id) \
        .order_by(Payment.year.desc(), Payment.month.desc()).all()

    history = [
        {
            "month": p.month,
            "year": p.year,
            "amount": float(p.amount) if p.amount else 0.0,
            "status": p.status,  # вместо p.paid или p.is_paid
            "paid": p.status == "paid"
        }
        for p in payments
    ]
    return jsonify(history)


@app.route('/payments/pay', methods=['POST'])
@login_required
@role_required('trainer')
def pay_payment():
    player_id = request.form.get('player_id', type=int)
    month = request.form.get('month', type=int)
    year = request.form.get('year', type=int)
    amount = request.form.get('amount', type=float)

    if not (player_id and month and year):
        return jsonify({"ok": False, "error": "Missing data"}), 400

    # Търсим съществуващ запис или създаваме нов
    payment = Payment.query.filter_by(player_id=player_id, month=month, year=year).first()
    if not payment:
        payment = Payment(
            player_id=player_id,
            month=month,
            year=year,
            amount=amount,
            status='paid'
        )
        db.session.add(payment)
    else:
        payment.amount = amount
        payment.status = 'paid'

    db.session.commit()

    # Известяване
    player = payment.player
    date_str = f"{month:02d}.{year}"
    msg = f"✅ Плащането за {date_str} е получено. Сума: {amount:.2f} лв."
    if player.email:
        send_email(player.email, "Потвърждение за плащане", msg)
    if player.parent_telegram_id:
        send_telegram(player.parent_telegram_id, msg)

    return jsonify({"ok": True})






# -------------------- Initialize App --------------------
def init_app():
    """Initialize the application - create tables and admin user."""
    with app.app_context():
        ensure_templates()
        db.create_all()
        
        # Create admin user if it doesn't exist
        admin_user = User.query.filter_by(username='admin').first()
        if not admin_user:
            admin_user = User(username='admin', role='admin')
            admin_user.set_password('admin123')
            db.session.add(admin_user)
            db.session.commit()
            logger.info("Admin user created")
        
        # Set webhook if in production
        maybe_set_webhook()

        # Schedule background jobs
        try:
            # Monthly reminders (hourly check)
            scheduler.add_job(send_monthly_reminders, 'interval', minutes=60, id='monthly_reminders', replace_existing=True)

            # Daily materialization wrapper (defers import until run time)
            def _materialize_wrapper():
                try:
                    from datetime import date, timedelta
                    from app_v3 import materialize_recurring_slots
                    materialize_recurring_slots(date.today(), date.today() + timedelta(days=7))
                except Exception:
                    logger.exception('materialize wrapper failed')

            scheduler.add_job(_materialize_wrapper, 'cron', hour=3, minute=0, id='materialize_slots_daily', replace_existing=True)

            if not scheduler.running:
                scheduler.start()
        except Exception:
            logger.exception('Failed to start scheduler jobs')

# Initialize app when imported
init_app()

# -------------------- Run --------------------
if __name__ == '__main__':
    init_app()
    app.run(host='0.0.0.0', port=5000, debug=True)

# ---------- Attendance Statistics by Player ----------
@app.route('/stats/attendance_by_player')
@login_required
@role_required('trainer')
def stats_attendance_by_player():
    # Взимаме статистика за присъствие по състезатели
    attendance_stats = db.session.query(
        Player.full_name,
        Player.team_id,
        db.func.count(Attendance.id).label("total_sessions"),
        db.func.sum(case((Attendance.status == 'present', 1), else_=0)).label("present_sessions")
    ).join(Attendance, Attendance.player_id == Player.id) \
     .group_by(Player.id, Player.full_name, Player.team_id) \
     .order_by(Player.full_name).all()

    # Изчисляваме процентите
    stats_list = []
    for stat in attendance_stats:
        percent = round((stat.present_sessions / stat.total_sessions) * 100, 1) if stat.total_sessions > 0 else 0
        team_name = Team.query.get(stat.team_id).name if stat.team_id else "Без отбор"
        stats_list.append({
            "full_name": stat.full_name,
            "team_name": team_name,
            "total_sessions": stat.total_sessions,
            "present_sessions": stat.present_sessions,
            "absent_sessions": stat.total_sessions - stat.present_sessions,
            "percent": percent
        })

    return render_template('attendance_stats.html', stats=stats_list)

@app.route('/stats/attendance_by_player_csv')
@login_required
@role_required('trainer')
def stats_attendance_by_player_csv():
    # Взимаме статистика за присъствие по състезатели
    attendance_stats = db.session.query(
        Player.full_name,
        Player.team_id,
        db.func.count(Attendance.id).label("total_sessions"),
        db.func.sum(case((Attendance.status == 'present', 1), else_=0)).label("present_sessions")
    ).join(Attendance, Attendance.player_id == Player.id) \
     .group_by(Player.id, Player.full_name, Player.team_id) \
     .order_by(Player.full_name).all()

    # Генерираме CSV
    output = StringIO()
    writer = csv.writer(output)

    # Заглавен ред
    header = ["Състезател", "Отбор", "Общо тренировки", "Присъствал", "Отсъствал", "Процент присъствие"]
    writer.writerow(header)

    # Редове за състезатели
    for stat in attendance_stats:
        percent = round((stat.present_sessions / stat.total_sessions) * 100, 1) if stat.total_sessions > 0 else 0
        team_name = Team.query.get(stat.team_id).name if stat.team_id else "Без отбор"
        row = [
            stat.full_name,
            team_name,
            stat.total_sessions,
            stat.present_sessions,
            stat.total_sessions - stat.present_sessions,
            f"{percent}%"
        ]
        writer.writerow(row)

    # Връщаме като отговор за сваляне
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=attendance_by_player.csv"}
    )

@app.route('/stats/attendance_by_player_excel')
@login_required
@role_required('trainer')
def stats_attendance_by_player_excel():
    # Взимаме статистика за присъствие по състезатели
    attendance_stats = db.session.query(
        Player.full_name,
        Player.team_id,
        db.func.count(Attendance.id).label("total_sessions"),
        db.func.sum(case((Attendance.status == 'present', 1), else_=0)).label("present_sessions")
    ).join(Attendance, Attendance.player_id == Player.id) \
     .group_by(Player.id, Player.full_name, Player.team_id) \
     .order_by(Player.full_name).all()

    # Създаваме Excel файл
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Статистика присъствие"

    # Стилове
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")

    # Заглавен ред
    headers = ["Състезател", "Отбор", "Общо тренировки", "Присъствал", "Отсъствал", "Процент присъствие"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    # Данни
    for row, stat in enumerate(attendance_stats, 2):
        percent = round((stat.present_sessions / stat.total_sessions) * 100, 1) if stat.total_sessions > 0 else 0
        team_name = Team.query.get(stat.team_id).name if stat.team_id else "Без отбор"
        
        ws.cell(row=row, column=1, value=stat.full_name)
        ws.cell(row=row, column=2, value=team_name)
        ws.cell(row=row, column=3, value=stat.total_sessions)
        ws.cell(row=row, column=4, value=stat.present_sessions)
        ws.cell(row=row, column=5, value=stat.total_sessions - stat.present_sessions)
        ws.cell(row=row, column=6, value=f"{percent}%")

    # Автоматично разширяване на колоните
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Запазваме файла
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=attendance_by_player.xlsx"}
    )

# ---------- Admin ----------

# ---------- Team Player Management ----------
@app.route('/teams/<int:team_id>/players')
@login_required
@role_required('trainer')
def team_players(team_id):
    team = Team.query.get_or_404(team_id)
    players = Player.query.filter_by(team_id=team_id).order_by(Player.full_name).all()
    all_teams = Team.query.order_by(Team.name).all()
    all_players = Player.query.order_by(Player.full_name).all()
    
    return render_template('team_players.html', team=team, players=players, all_teams=all_teams, all_players=all_players)

@app.route('/teams/<int:team_id>/players/move', methods=['POST'])
@login_required
@role_required('trainer')
def move_player_to_team():
    player_id = request.form.get('player_id', type=int)
    new_team_id = request.form.get('new_team_id', type=int)
    
    if not player_id or not new_team_id:
        flash('Липсват данни', 'error')
        return redirect(request.referrer or url_for('teams'))
    
    player = Player.query.get_or_404(player_id)
    new_team = Team.query.get_or_404(new_team_id)
    
    old_team_name = player.team.name if player.team else "Без отбор"
    player.team_id = new_team_id
    db.session.commit()
    
    flash(f'Състезателят {player.full_name} е преместен от {old_team_name} в {new_team.name}', 'success')
    return redirect(request.referrer or url_for('teams'))

@app.route('/teams/<int:team_id>/players/remove', methods=['POST'])
@login_required
@role_required('trainer')
def remove_player_from_team():
    player_id = request.form.get('player_id', type=int)
    
    if not player_id:
        flash('Липсват данни', 'error')
        return redirect(request.referrer or url_for('teams'))
    
    player = Player.query.get_or_404(player_id)
    team_name = player.team.name if player.team else "Без отбор"
    player.team_id = None
    db.session.commit()
    
    flash(f'Състезателят {player.full_name} е премахнат от {team_name}', 'success')
    return redirect(request.referrer or url_for('teams'))

@app.route('/teams/<int:team_id>/players/add', methods=['POST'])
@login_required
@role_required('trainer')
def add_player_to_team():
    team_id = request.form.get('team_id', type=int)
    player_id = request.form.get('player_id', type=int)
    
    if not team_id or not player_id:
        flash('Липсват данни', 'error')
        return redirect(request.referrer or url_for('teams'))
    
    player = Player.query.get_or_404(player_id)
    team = Team.query.get_or_404(team_id)
    
    old_team_name = player.team.name if player.team else "Без отбор"
    player.team_id = team_id
    db.session.commit()
    
    flash(f'Състезателят {player.full_name} е добавен в {team.name}', 'success')
    return redirect(request.referrer or url_for('teams'))

# ---------- Admin ----------

@app.route('/calendar')
@login_required
def calendar_view():
    return render_template('calendar.html')

@app.route('/api/calendar/events')
@login_required
def api_calendar_events():
    start = request.args.get('start')
    end = request.args.get('end')

    # Determine range and materialize sessions from recurring slots automatically
    try:
        range_start = datetime.fromisoformat(start[:10]).date() if start else date.today()
    except Exception:
        range_start = date.today()
    try:
        range_end = datetime.fromisoformat(end[:10]).date() if end else (range_start + timedelta(days=14))
    except Exception:
        range_end = range_start + timedelta(days=14)

    # Create TrainingSession from active season slots for the requested period
    try:
        materialize_recurring_slots(range_start, range_end)
    except Exception:
        logger.exception('materialize_recurring_slots during /api/calendar/events failed')

    q = TrainingSession.query
    q = q.filter(TrainingSession.date >= range_start, TrainingSession.date <= range_end)

    sessions = q.order_by(TrainingSession.date.asc()).all()

    def guess_venue(notes: str) -> str:
        if not notes:
            return ''
        text = notes.strip().upper()
        if 'НУПИ' in text:
            return 'НУПИ'
        if 'ЧАВДАР' in text:
            return 'ЧАВДАР'
        if 'СТАДИОН' in text:
            return 'СТАДИОН'
        return ''

    # използваме централизираната цветова функция
    def team_color_for(name: str) -> str:
        return team_color_for_name(name)

    events = []
    for s in sessions:
        start_time = (s.start_time or '08:00')
        end_time = (s.end_time or '09:00')
        start_iso = f"{s.date.isoformat()}T{start_time}:00"
        end_iso = f"{s.date.isoformat()}T{end_time}:00"
        team_name = Team.query.get(s.team_id).name if s.team_id else '—'
        color = team_color_for(team_name)
        events.append({
            'id': f'tr-{s.id}',
            'title': team_name,
            'start': start_iso,
            'end': end_iso,
            'venue': guess_venue(s.notes or ''),
            'edit_url': url_for('edit_training', training_id=s.id),
            'team_color': color
        })

    # We no longer render raw recurring slots here because they are materialized above.

    return jsonify(events)

# -------- Season & Recurring slots (basic admin) --------
@app.route('/admin/seasons', methods=['GET','POST'])
@role_required('admin')
def manage_seasons():
    if request.method == 'POST':
        name = request.form.get('name')
        start = request.form.get('start')
        end = request.form.get('end')
        active = bool(request.form.get('is_active'))
        s = Season(name=name,
                   start_date=datetime.fromisoformat(start).date() if start else None,
                   end_date=datetime.fromisoformat(end).date() if end else None,
                   is_active=active)
        if active:
            Season.query.update({Season.is_active: False})
            s.is_active = True
        db.session.add(s); db.session.commit()
        flash('Сезонът е записан','success')
        return redirect(url_for('manage_seasons'))
    seasons = Season.query.order_by(Season.id.desc()).all()
    return render_template('seasons.html', seasons=seasons)

@app.route('/admin/slots', methods=['GET','POST'])
@role_required('admin')
def manage_recurring_slots():
    seasons = Season.query.order_by(Season.id.desc()).all()
    teams = Team.query.order_by(Team.name).all()
    if request.method == 'POST':
        season_id = request.form.get('season_id', type=int)
        team_id = request.form.get('team_id', type=int)
        weekday = request.form.get('weekday', type=int)
        start_time = request.form.get('start_time')
        end_time = request.form.get('end_time')
        venue = request.form.get('venue')
        title = request.form.get('title')
        slot = RecurringSlot(season_id=season_id, team_id=team_id, weekday=weekday,
                             start_time=start_time, end_time=end_time, venue=venue, title=title)
        db.session.add(slot); db.session.commit()
        flash('Слотът е добавен','success')
        return redirect(url_for('manage_recurring_slots'))
    slots = RecurringSlot.query.order_by(RecurringSlot.weekday, RecurringSlot.start_time).all()
    return render_template('slots.html', seasons=seasons, teams=teams, slots=slots)

@app.route('/admin/slots/<int:slot_id>/delete', methods=['POST'])
@role_required('admin')
def delete_recurring_slot(slot_id):
    slot = RecurringSlot.query.get_or_404(slot_id)
    db.session.delete(slot)
    db.session.commit()
    flash('Слотът е изтрит','success')
    return redirect(url_for('manage_recurring_slots'))

@app.route('/admin/slots/delete_all', methods=['POST'])
@role_required('admin')
def delete_all_trainings():
    """Изтриване на всички TrainingSession (чистене на стар график)."""
    count = TrainingSession.query.count()
    Attendance.query.delete()
    db.session.execute(db.delete(TrainingSession))
    db.session.commit()
    flash(f'Изтрити тренировки: {count}','success')
    return redirect(url_for('trainings'))

@app.route('/admin/seed_summer_2024_2025', methods=['POST'])
@role_required('admin')
def seed_summer_season():
    # Create or get season
    name = 'ЛЕТЕН 2024/2025'
    season = Season.query.filter_by(name=name).first()
    if not season:
        season = Season(name=name, is_active=True)
        # deactivate others
        Season.query.update({Season.is_active: False})
        db.session.add(season)
        db.session.commit()
    else:
        # set active
        Season.query.update({Season.is_active: False})
        season.is_active = True
        db.session.commit()

    def team_by_name(n):
        return Team.query.filter_by(name=n).first()

    slots = []
    # Helper to append slot
    def add_slot(weekday, start_time, end_time, venue, team_label):
        # Нормализирани етикети според графика
        name = _normalize_team_label(team_label)
        t = team_by_name(name)
        team_id = t.id if t else None
        slots.append(RecurringSlot(
            season_id=season.id,
            team_id=team_id,
            weekday=weekday,
            start_time=start_time,
            end_time=end_time,
            venue=venue,
            title=name
        ))

    # Понеделник (0)
    add_slot(0, '08:00', '09:00', 'СТАДИОН', 'U-12 Ж')
    add_slot(0, '08:00', '09:00', 'СТАДИОН', 'U-12 М')
    add_slot(0, '10:00', '11:30', 'ЧАВДАР', 'U-18 Ж')
    add_slot(0, '16:00', '17:30', 'НУПИ', 'U-12 Ж')
    add_slot(0, '17:30', '19:00', 'НУПИ', 'U-12 М')
    add_slot(0, '19:30', '21:15', 'ЧАВДАР', 'U-18 М')

    # Вторник (1)
    add_slot(1, '10:00', '11:30', 'ЧАВДАР', 'U-18 Ж')
    add_slot(1, '11:00', '12:30', 'ЧАВДАР', 'U-18 М')

    # Сряда (2)
    add_slot(2, '08:00', '09:00', 'СТАДИОН', 'U-12 Ж')
    add_slot(2, '08:00', '09:00', 'СТАДИОН', 'U-12 М')
    add_slot(2, '10:00', '11:30', 'ЧАВДАР', 'U-18 Ж')
    add_slot(2, '16:00', '17:30', 'НУПИ', 'U-12 Ж')
    add_slot(2, '17:30', '19:00', 'НУПИ', 'U-12 М')
    add_slot(2, '19:30', '21:30', 'ЧАВДАР', 'U-18 М')

    # Четвъртък (3)
    add_slot(3, '08:00', '09:00', 'СТАДИОН', 'U-18 М')
    add_slot(3, '10:00', '11:30', 'НУПИ', 'U-12 Ж')
    add_slot(3, '11:30', '13:00', 'НУПИ', 'U-12 Ж')

    # Петък (4)
    add_slot(4, '08:30', '10:00', 'НУПИ', 'U-18 Ж')
    add_slot(4, '10:00', '11:30', 'НУПИ', 'U-12 М')
    add_slot(4, '10:00', '12:00', 'НУПИ', 'U-12 М')

    # Събота (5)
    add_slot(5, '19:30', '21:15', 'ЧАВДАР', 'U-18 М')

    # Save (clear previous slots of the season)
    RecurringSlot.query.filter_by(season_id=season.id).delete()
    db.session.add_all(slots)
    db.session.commit()
    flash('Летният график е въведен и сезонът е активен.','success')
    return redirect(url_for('manage_recurring_slots'))

# --- Seed from image (U-12 Ж/М, U-18 Ж/М) ---
@app.route('/admin/seed_from_image', methods=['POST'])
@role_required('admin')
def seed_from_image():
    name = 'ЛЕТЕН Сезон 2024/2025'
    # Activate or create season
    season = Season.query.filter_by(name=name).first()
    if not season:
        Season.query.update({Season.is_active: False})
        season = Season(name=name, is_active=True)
        db.session.add(season); db.session.commit()
    else:
        Season.query.update({Season.is_active: False})
        season.is_active = True; db.session.commit()

    ensure_schedule_teams()

    def team_id_by(name):
        t = Team.query.filter_by(name=name).first()
        return t.id if t else None

    def add(weekday, start_time, end_time, venue, title):
        return RecurringSlot(
            season_id=season.id,
            team_id=team_id_by(title),
            weekday=weekday,
            start_time=start_time,
            end_time=end_time,
            venue=venue,
            title=title
        )

    slots = [
        # Понеделник (0)
        add(0, '08:00', '09:00', 'СТАДИОН', 'U-12 Ж'),
        add(0, '08:00', '09:00', 'СТАДИОН', 'U-12 М'),
        add(0, '10:00', '11:30', 'ЧАВДАР', 'U-18 Ж'),
        add(0, '16:00', '17:30', 'НУПИ', 'U-12 Ж'),
        add(0, '17:30', '19:00', 'НУПИ', 'U-12 Ж'),
        add(0, '19:30', '21:15', 'ЧАВДАР', 'U-18 М'),
        # Вторник (1)
        add(1, '10:00', '11:30', 'ЧАВДАР', 'U-18 Ж'),
        add(1, '11:00', '12:30', 'ЧАВДАР', 'U-18 М'),
        # Сряда (2)
        add(2, '08:00', '09:00', 'СТАДИОН', 'U-12 Ж'),
        add(2, '08:00', '09:00', 'СТАДИОН', 'U-12 М'),
        add(2, '10:00', '11:30', 'ЧАВДАР', 'U-18 Ж'),
        add(2, '16:00', '17:30', 'НУПИ', 'U-12 Ж'),
        add(2, '17:30', '19:00', 'НУПИ', 'U-12 М'),
        add(2, '19:30', '21:30', 'ЧАВДАР', 'U-18 М'),
        # Четвъртък (3)
        add(3, '08:00', '09:00', 'СТАДИОН', 'U-18 М'),
        add(3, '10:00', '11:30', 'НУПИ', 'U-12 Ж'),
        add(3, '11:30', '13:00', 'НУПИ', 'U-12 Ж'),
        # Петък (4)
        add(4, '08:30', '10:00', 'НУПИ', 'U-18 Ж'),
        add(4, '10:00', '11:30', 'НУПИ', 'U-12 М'),
        add(4, '10:00', '12:00', 'НУПИ', 'U-12 М'),
        # Събота (5)
        add(5, '19:30', '21:15', 'ЧАВДАР', 'U-18 М'),
    ]

    # Clear previous slots of the season and all future materialized trainings
    RecurringSlot.query.filter_by(season_id=season.id).delete()
    db.session.add_all(slots)
    db.session.commit()

    # По желание: изчистваме всички тренировки (стар график)
    Attendance.query.delete()
    db.session.execute(db.delete(TrainingSession))
    db.session.commit()

    flash('Графикът от снимката е въведен. Старите тренировки са изтрити.','success')
    return redirect(url_for('manage_recurring_slots'))

# -------- Helpers: Teams normalization and materialization of slots --------

SCHEDULE_TEAM_NAMES = [
    'U-12 Ж',
    'U-12 М',
    'U-18 Ж',
    'U-18 М',
    'Старша',
]

def ensure_schedule_teams():
    """Ensure teams matching schedule names exist. Optionally rename default U* teams."""
    existing_by_name = {t.name: t for t in Team.query.all()}

    # Try to normalize some demo names into the new scheme
    rename_map = {}
    for t in Team.query.all():
        lname = (t.name or '').lower()
        if ('u12' in lname and 'girls' in lname) or 'момичета до 12' in lname:
            rename_map[t.id] = 'U-12 Ж'
        elif ('u12' in lname and 'boys' in lname) or ('момчета' in lname and '12' in lname):
            rename_map[t.id] = 'U-12 М'
        elif ('u18' in lname and 'girls' in lname) or ('момичета' in lname and '18' in lname):
            rename_map[t.id] = 'U-18 Ж'
        elif ('u18' in lname and 'boys' in lname) or ('момчета' in lname and '18' in lname) or ('мъже' in lname):
            rename_map[t.id] = 'U-18 М'
        elif 'senior' in lname or 'старша' in lname:
            rename_map[t.id] = 'Старша'
    # Apply renames if target name not already taken
    for team_id, new_name in rename_map.items():
        if new_name not in existing_by_name:
            t = Team.query.get(team_id)
            if t:
                t.name = new_name
                existing_by_name[new_name] = t
    # Ensure all schedule teams exist
    for n in SCHEDULE_TEAM_NAMES:
        if n not in existing_by_name:
            db.session.add(Team(name=n))
    db.session.commit()

def resolve_team_id_for_slot(slot: RecurringSlot) -> Optional[int]:
    if slot.team_id:
        return slot.team_id
    # try match by title
    title = (slot.title or '').strip()
    if not title:
        return None
    team = Team.query.filter_by(name=title).first()
    if not team:
        team = Team(name=title)
        db.session.add(team)
        db.session.commit()
    return team.id

def materialize_recurring_slots(range_start: date, range_end: date) -> int:
    """Create TrainingSession rows from active season RecurringSlot in the given range.
    Returns number of created sessions.
    """
    created = 0
    active_season = Season.query.filter_by(is_active=True).first()
    if not active_season:
        return 0
    ensure_schedule_teams()
    slots = RecurringSlot.query.filter_by(season_id=active_season.id).all()
    d = range_start
    while d <= range_end:
        for slot in slots:
            if d.weekday() != slot.weekday:
                continue
            team_id = resolve_team_id_for_slot(slot)
            # prevent duplicates by (team_id, date, start_time)
            exists = TrainingSession.query.filter_by(
                team_id=team_id, date=d, start_time=slot.start_time
            ).first()
            if exists:
                continue
            ts = TrainingSession(
                team_id=team_id,
                date=d,
                start_time=slot.start_time,
                end_time=slot.end_time,
                notes=slot.venue or slot.title or ''
            )
            db.session.add(ts)
            created += 1
        d = d + timedelta(days=1)
    if created:
        db.session.commit()
    return created

def scheduled_materialize_upcoming():
    try:
        with app.app_context():
            materialize_recurring_slots(date.today(), date.today() + timedelta(days=7))
    except Exception:
        logger.exception('scheduled_materialize_upcoming failed')

@app.route('/admin/materialize_slots', methods=['POST'])
@role_required('admin')
def admin_materialize_slots():
    """Materialize slots to sessions in a given range (defaults next 14 days)."""
    start_str = request.form.get('start')
    end_str = request.form.get('end')
    try:
        rs = datetime.fromisoformat(start_str).date() if start_str else date.today()
        re_ = datetime.fromisoformat(end_str).date() if end_str else (rs + timedelta(days=14))
    except Exception:
        rs = date.today()
        re_ = rs + timedelta(days=14)
    created = materialize_recurring_slots(rs, re_)
    flash(f'Създадени тренировки от слотове: {created}', 'success')
    return redirect(url_for('trainings'))

@app.route('/admin/materialize_month', methods=['POST'])
@role_required('admin')
def admin_materialize_month():
    """Материализира всички слотове за подадения месец (year, month)."""
    y = request.form.get('year', type=int)
    m = request.form.get('month', type=int)
    if not y or not m:
        flash('Липсват year/month', 'error')
        return redirect(url_for('trainings'))
    # първи и последен ден на месеца
    start = date(y, m, 1)
    if m == 12:
        end = date(y + 1, 1, 1) - timedelta(days=1)
    else:
        end = date(y, m + 1, 1) - timedelta(days=1)
    created = materialize_recurring_slots(start, end)
    flash(f'Материализирани тренировки за {m:02d}.{y}: {created}', 'success')
    return redirect(url_for('trainings', year=y, month=m))
